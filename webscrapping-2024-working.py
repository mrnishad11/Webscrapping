import concurrent
import datetime
import logging
import urllib
from concurrent.futures import ALL_COMPLETED
import openpyxl
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import pandas as pd

from nsedt import utils
from nsedt.resources import constants as cns
from nsedt.utils import data_format
from nsedt import equity as eq
from datetime import date
log = logging.getLogger("root")


def get_companyinfo(
    symbol: str,
    response_type: str = "panda_df",
):
    params = {}
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.EQUITY_INFO

    params["symbol"] = symbol

    url = base_url + event_api + urllib.parse.urlencode(params)
    data = utils.fetch_url(
        url,
        cookies,
        key=None,
        response_type=response_type,
    )

    return data


def get_marketstatus(
    response_type: str = "panda_df",
):
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.MARKETSTATUS

    url = base_url + event_api
    data = utils.fetch_url(
        url,
        cookies,
        key="marketState",
        response_type=response_type,
    )

    return data


def get_price(
    start_date,
    end_date,
    symbol=None,
    input_type="stock",
    series="EQ",
):
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    price_api = cns.EQUITY_PRICE_HISTORY
    url_list = []

    # set the window size to one year
    window_size = datetime.timedelta(days=cns.WINDOW_SIZE)

    start_date, end_date = utils.check_nd_convert(start_date, end_date)

    current_window_start = start_date
    while current_window_start < end_date:
        current_window_end = current_window_start + window_size

        # check if the current window extends beyond the end_date
        current_window_end = min(current_window_end, end_date)

        if input_type == "stock":
            params = {
                "symbol": symbol,
                "from": current_window_start.strftime("%d-%m-%Y"),
                "to": current_window_end.strftime("%d-%m-%Y"),
                "dataType": "priceVolumeDeliverable",
                "series": series,
            }
            url = base_url + price_api + urllib.parse.urlencode(params)
            url_list.append(url)

        # move the window start to the next day after the current window end
        current_window_start = current_window_end + datetime.timedelta(days=1)

    result = pd.DataFrame()
    with concurrent.futures.ThreadPoolExecutor(max_workers=cns.MAX_WORKERS) as executor:
        future_to_url = {
            executor.submit(utils.fetch_url, url, cookies, "data"): url
            for url in url_list
        }
        concurrent.futures.wait(future_to_url, return_when=ALL_COMPLETED)
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                dataframe = future.result()
                result = pd.concat([result, dataframe])
            except Exception as exc:
                logging.error("%s got exception: %s. Please try again later.", url, exc)
                raise exc
    return data_format.price(result)


def get_corpinfo(
    start_date,
    end_date,
    symbol=None,
    response_type="panda_df",
):
    cookies = utils.get_cookies()
    params = {
        "symbol": symbol,
        "from_date": start_date,
        "to_date": end_date,
        "index": "equities",
    }
    base_url = cns.BASE_URL
    price_api = cns.EQUITY_CORPINFO
    url = base_url + price_api + urllib.parse.urlencode(params)

    data = utils.fetch_url(
        url,
        cookies,
        key=None,
        response_type=response_type,
    )

    return data


def get_event(
    start_date=None,
    end_date=None,
    index="equities",
):
    params = {}
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.EQUITY_EVENT

    params["index"] = index
    if start_date is not None:
        params["from_date"] = start_date
    if end_date is not None:
        params["to_date"] = end_date

    url = base_url + event_api + urllib.parse.urlencode(params)
    return utils.fetch_url(url, cookies)


def get_chartdata(
    symbol,
    preopen=False,
    response_type="panda_df",
):
    params = {}
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.EQUITY_CHART
    try:
        identifier = get_companyinfo(
            symbol,
            response_type="json",
        )[
            "info"
        ]["identifier"]

    except KeyError:
        return f"Invalid symbol name: {symbol}"

    params["index"] = identifier
    if preopen:
        params["preopen"] = "true"

    url = base_url + event_api + urllib.parse.urlencode(params)

    data = utils.fetch_url(
        url,
        cookies,
        key="grapthData",
        response_type=response_type,
    )
    if response_type == "panda_df":
        data_frame = data.rename(
            columns={
                0: "timestamp_milliseconds",
                1: "price",
            }
        )
        data_frame["datetime"] = pd.to_datetime(
            data_frame["timestamp_milliseconds"], unit="ms"
        )
        return data_frame
    return data


def get_symbols_list():

    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.EQUITY_LIST

    url = base_url + event_api
    data = utils.fetch_url(url, cookies)
    f_dict = data.to_dict()
    eq_list = []
    for i in range(len(f_dict["data"])):
        eq_list.append(f_dict["data"][i]["metadata"]["symbol"])

    return eq_list


def get_asm_list(asm_type="both") -> list:
    cookies = utils.get_cookies()
    base_url = cns.BASE_URL
    event_api = cns.ASM_LIST

    url = base_url + event_api
    data = utils.fetch_url(url, cookies)
    _data = data.to_dict()

    if asm_type ==  "both":
        return _data
    if asm_type == "longterm":
        return _data.get("longterm").get("data")
    if asm_type == "shortterm":
        return _data.get("shortterm").get("data")
    return ["possible values are both,longterm,shortterm"]
start_date = date(2023, 1, 1)
end_date = date(2024, 4, 10)
stock_name="BLUESTARCO"
stock_data=eq.get_price(start_date, end_date, symbol=stock_name)

stock_data.to_excel(f'{stock_name}.xlsx')
stockname=stock_name+'.xlsx'
wb = openpyxl.load_workbook(stockname)
ws=wb.active    

ws['R1']='ANALYSIS'
ws['R1'].font = Font(bold=True,color="FF0000")
#DATE
for rows in range(1,len(stock_data)):
    paste='S'+str(rows)
    copy='B'+str(rows)
    ws[paste].value=ws[copy].value
#VWAP
for rows in range(1,len(stock_data)):
    paste='T'+str(rows)
    copy='L'+str(rows)
    ws[paste].value=ws[copy].value
#change in price
ws['U1']='~PRICE'
ws['U1'].font = Font(bold=True)
for rows in range(2,len(stock_data)):
    copy1='F'+str(rows)
    copy2='F'+str(rows+1)
    paste='U'+str(rows+1)
    if ws[copy1].value==None or ws[copy2].value==None:
        break
    ws[paste]=((ws[copy2].value-ws[copy1].value)/ws[copy2].value)
    ws[paste].number_format=FORMAT_PERCENTAGE_00
   

#DELIVERY IN CRORE
ws['X1']='DELIVERY'
ws['X1'].font = Font(bold=True)
for rows in range(2,len(stock_data)):
    copy1='M'+str(rows)
    copy2='N'+str(rows)
    paste='X'+str(rows)
    if ws[copy1].value==None or ws[copy2].value==None:
        break
    ws[paste].value=(ws[copy1].value*ws[copy2].value)/10000000
    ws[paste].number_format='#,##0.0000'
#5 day Average
ws['Y1']='5 DAY AVG'
ws['Y1'].font = Font(bold=True)
for rows in range(2,len(stock_data)):
    copy1='X'+str(rows)
    copy2='X'+str(rows+1)
    copy3='X'+str(rows+2)
    copy4='X'+str(rows+3)
    copy5='X'+str(rows+4)
    paste='Y'+str(rows+4)
    if (ws[copy1].value!=None and ws[copy2].value!=None) and (ws[copy3].value!=None and ws[copy4].value!=None) and ws[copy5].value!=None:
        ws[paste].value=(ws[copy1].value+ws[copy2].value+ws[copy3].value+ws[copy4].value+ws[copy5].value)/5    
        ws[paste].number_format=FORMAT_PERCENTAGE_00
#Change in DELIVERY(~DELIVERY)
ws['Z1']='~DELIVERY'
ws['Z1'].font = Font(bold=True)
for rows in range(6,len(stock_data)):
    copy1='X'+str(rows)
    copy2='Y'+str(rows)
    paste='Z'+str(rows)
    if ws[copy1].value==None or ws[copy2].value==None:
        break
    ws[paste].value=ws[copy1].value/ws[copy2].value
    ws[paste].number_format=FORMAT_PERCENTAGE_00
	
    

wb.save(f'{stock_name} modified.xlsx')
print(f"Succeessfully Exported...!!!{stock_name}")
