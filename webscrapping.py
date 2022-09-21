import openpyxl
from nsepy import get_history
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
start_date=datetime(2022,5,1)
end_date=datetime(2022,8,25)
stock=['MUKANDLTD']
for i in stock:
    stock_data=get_history(symbol=i,start=start_date,end=end_date)
    stock_data.to_excel(f'{i}.xlsx')
    #print("Successfully Downloaded The File....!!!")
    stockname=i+'.xlsx'
    wb = openpyxl.load_workbook(stockname)
    ws=wb.active    

    ws['R1']='ANALYSIS'
    ws['R1'].font = Font(bold=True,color="FF0000")
#DATE
    for rows in range(1,180):
        paste='S'+str(rows)
        copy='A'+str(rows)
        ws[paste].value=ws[copy].value
#VWAP
    for rows in range(1,127):
        paste='T'+str(rows)
        copy='J'+str(rows)
        ws[paste].value=ws[copy].value
#change in price
    ws['U1']='~PRICE'
    ws['U1'].font = Font(bold=True)
    for rows in range(2,180):
        copy1='T'+str(rows)
        copy2='T'+str(rows+1)
        paste='U'+str(rows+1)
        if ws[copy1].value==None or ws[copy2].value==None:
            break
        ws[paste]=((ws[copy2].value-ws[copy1].value)/ws[copy2].value)
        ws[paste].number_format=FORMAT_PERCENTAGE_00
   

#DELIVERY IN CRORE
    ws['X1']='DELIVERY'
    ws['X1'].font = Font(bold=True)
    for rows in range(2,126):
        copy1='N'+str(rows)
        copy2='T'+str(rows)
        paste='X'+str(rows)
        if ws[copy1].value==None or ws[copy2].value==None:
            break
        ws[paste].value=(ws[copy1].value*ws[copy2].value)/10000000
        ws[paste].number_format='#,##0.0000'
#5 day Average
    ws['Y1']='5 DAY AVG'
    ws['Y1'].font = Font(bold=True)
    for rows in range(2,180):
        copy1='X'+str(rows)
        copy2='X'+str(rows+1)
        copy3='X'+str(rows+2)
        copy4='X'+str(rows+3)
        copy5='X'+str(rows+4)
        paste='Y'+str(rows+4)
        if (ws[copy1].value!=None and ws[copy2].value!=None) and (ws[copy3].value!=None and ws[copy4].value!=None) and ws[copy5].value!=None:
            ws[paste].value=(ws[copy1].value+ws[copy2].value+ws[copy3].value+ws[copy4].value+ws[copy5].value)/5    
            ws[paste].number_format=FORMAT_PERCENTAGE_00
#Change in DELIVERY(~DELIVERY)
    ws['Z1']='~DELIVERY'
    ws['Z1'].font = Font(bold=True)
    for rows in range(6,180):
        copy1='X'+str(rows)
        copy2='Y'+str(rows)
        paste='Z'+str(rows)
        if ws[copy1].value==None or ws[copy2].value==None:
            break
        ws[paste].value=ws[copy1].value/ws[copy2].value
        ws[paste].number_format=FORMAT_PERCENTAGE_00
	
    

    wb.save(f'{i} modified.xlsx')
    print(f"Succeessfully Exported...!!!{i}")

